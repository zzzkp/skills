#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from urllib.parse import quote

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "缺少依赖：openpyxl。请先执行：python -m pip install openpyxl"
    ) from exc


SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


@dataclass
class SheetDoc:
    name: str
    filename: str
    rows: list[list[str]]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return max((len(row) for row in self.rows), default=0)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.replace(microsecond=0).isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat()
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return text.strip()


def markdown_escape(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("\\", "\\\\").replace("|", "\\|")
    return text.replace("\n", "<br>")


def markdown_link_target(filename: str) -> str:
    return quote(filename.replace("\\", "/"), safe="/._-~")


def is_empty_row(row: list[str]) -> bool:
    return not any(cell.strip() for cell in row)


def fill_merged_values(sheet) -> list[list[str]]:
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    grid = [
        [
            normalize_text(sheet.cell(row=row, column=col).value)
            for col in range(1, max_col + 1)
        ]
        for row in range(1, max_row + 1)
    ]

    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col_range, max_row_range = merged_range.bounds
        value = normalize_text(sheet.cell(row=min_row, column=min_col).value)
        if not value:
            continue
        for row in range(min_row, max_row_range + 1):
            for col in range(min_col, max_col_range + 1):
                grid[row - 1][col - 1] = value

    return grid


def trim_empty_edges(rows: list[list[str]]) -> list[list[str]]:
    non_empty_rows = [index for index, row in enumerate(rows) if not is_empty_row(row)]
    if not non_empty_rows:
        return []

    start_row = non_empty_rows[0]
    end_row = non_empty_rows[-1]
    scoped_rows = rows[start_row : end_row + 1]

    non_empty_cols: list[int] = []
    max_width = max((len(row) for row in scoped_rows), default=0)
    for col_index in range(max_width):
        if any(
            col_index < len(row) and row[col_index].strip()
            for row in scoped_rows
        ):
            non_empty_cols.append(col_index)

    if not non_empty_cols:
        return []

    start_col = non_empty_cols[0]
    end_col = non_empty_cols[-1]
    return [
        [
            row[col_index] if col_index < len(row) else ""
            for col_index in range(start_col, end_col + 1)
        ]
        for row in scoped_rows
    ]


def column_label(index: int) -> str:
    label = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        label = chr(65 + remainder) + label
    return label


def pad_rows(rows: list[list[str]], width: int) -> list[list[str]]:
    return [row + [""] * (width - len(row)) for row in rows]


def build_headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        header = normalize_text(value) or column_label(index)
        count = used.get(header, 0) + 1
        used[header] = count
        headers.append(header if count == 1 else f"{header}_{count}")
    return headers


def render_markdown_table(rows: list[list[str]], first_row_header: bool) -> list[str]:
    width = max((len(row) for row in rows), default=0)
    if width == 0:
        return []

    padded_rows = pad_rows(rows, width)
    if first_row_header:
        headers = build_headers(padded_rows[0])
        body_rows = padded_rows[1:]
    else:
        headers = [column_label(index) for index in range(1, width + 1)]
        body_rows = padded_rows

    lines = [
        "| " + " | ".join(markdown_escape(header) for header in headers) + " |",
        "| " + " | ".join("---" for _header in headers) + " |",
    ]
    lines.extend(
        "| " + " | ".join(markdown_escape(cell) for cell in row) + " |"
        for row in body_rows
    )
    return lines


def sanitize_filename(value: str, used: set[str]) -> str:
    name = normalize_text(value) or "sheet"
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
    name = re.sub(r"\s+", "_", name)
    name = name.strip("._ ")
    if not name:
        name = "sheet"

    candidate = f"{name}.md"
    counter = 2
    while candidate.lower() in used:
        candidate = f"{name}_{counter}.md"
        counter += 1
    used.add(candidate.lower())
    return candidate


def render_sheet_md(sheet_doc: SheetDoc, source: Path, first_row_header: bool) -> str:
    lines = [
        f"# {markdown_escape(sheet_doc.name)}",
        "",
        "## 基本信息",
        "",
        "| 项目 | 内容 |",
        "|---|---|",
        f"| 来源文件 | `{markdown_escape(source.name)}` |",
        f"| Sheet | {markdown_escape(sheet_doc.name)} |",
        f"| 行数 | {sheet_doc.row_count} |",
        f"| 列数 | {sheet_doc.column_count} |",
        "",
        "## 数据",
        "",
    ]
    lines.extend(render_markdown_table(sheet_doc.rows, first_row_header=first_row_header))
    lines.append("")
    return "\n".join(lines)


def render_readme_md(sheet_docs: list[SheetDoc], source: Path) -> str:
    lines = [
        "# Excel 转 Markdown 总览",
        "",
        f"- 来源文件：`{source.name}`",
        f"- Sheet 数量：{len(sheet_docs)}",
        "",
        "| 序号 | Sheet | 行数 | 列数 | 文档 |",
        "|---:|---|---:|---:|---|",
    ]
    for index, sheet_doc in enumerate(sheet_docs, start=1):
        lines.append(
            "| "
            + " | ".join(
                [
                    str(index),
                    markdown_escape(sheet_doc.name),
                    str(sheet_doc.row_count),
                    str(sheet_doc.column_count),
                    f"[{markdown_escape(sheet_doc.filename)}]({markdown_link_target(sheet_doc.filename)})",
                ]
            )
            + " |"
        )
    lines.append("")
    return "\n".join(lines)


def write_text(path: Path, content: str, overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(f"拒绝覆盖已有文件：{path}")
    path.write_text(content, encoding="utf-8", newline="\n")


def convert(
    excel_path: Path,
    output_dir: Path,
    overwrite: bool,
    first_row_header: bool,
) -> list[SheetDoc]:
    workbook = load_workbook(excel_path, data_only=True)
    used_filenames = {"readme.md"}
    sheet_docs: list[SheetDoc] = []

    for sheet in workbook.worksheets:
        rows = trim_empty_edges(fill_merged_values(sheet))
        if not rows:
            continue
        filename = sanitize_filename(sheet.title, used_filenames)
        sheet_docs.append(SheetDoc(name=sheet.title, filename=filename, rows=rows))

    if not sheet_docs:
        raise RuntimeError("未检测到非空 Sheet。")

    output_dir.mkdir(parents=True, exist_ok=True)
    for sheet_doc in sheet_docs:
        write_text(
            output_dir / sheet_doc.filename,
            render_sheet_md(sheet_doc, excel_path, first_row_header=first_row_header),
            overwrite=overwrite,
        )

    write_text(output_dir / "README.md", render_readme_md(sheet_docs, excel_path), overwrite=overwrite)
    return sheet_docs


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="将 Excel 工作簿按 Sheet 转换为 Markdown。")
    parser.add_argument("excel", type=Path, help="Excel 文件路径。")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="输出目录。默认在 Excel 同级目录生成与源文件同名的目录。",
    )
    parser.add_argument(
        "--first-row-header",
        action="store_true",
        help="将每个 Sheet 的第一行作为 Markdown 表头；默认生成 A/B/C 列头并保留所有行。",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="覆盖输出目录中已存在的 Markdown 文件。",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    excel_path = args.excel.resolve()
    if not excel_path.exists():
        print(f"Excel 文件不存在：{excel_path}", file=sys.stderr)
        return 2
    if excel_path.suffix.lower() not in SUPPORTED_SUFFIXES:
        supported = "、".join(sorted(SUPPORTED_SUFFIXES))
        print(f"不支持的 Excel 类型：{excel_path.suffix}。支持：{supported}", file=sys.stderr)
        return 2

    output_dir = args.out or excel_path.with_name(excel_path.stem)
    try:
        sheet_docs = convert(
            excel_path,
            output_dir.resolve(),
            overwrite=args.overwrite,
            first_row_header=args.first_row_header,
        )
    except (FileExistsError, RuntimeError) as exc:
        print(str(exc), file=sys.stderr)
        return 1

    print(f"已生成 {len(sheet_docs)} 个 Sheet 文档：{output_dir.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
