#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install openpyxl"
    ) from exc


SKIP_SHEET_KEYWORDS = (
    "封面",
    "目录",
    "说明",
    "变更",
    "版本",
    "修订",
    "更新记录",
    "索引",
)

FIELD_HEADER_ALIASES = {
    "index": {
        "序号",
        "编号",
        "no",
        "num",
        "number",
        "idx",
    },
    "name": {
        "字段",
        "字段名",
        "字段名称",
        "列名",
        "属性名",
        "英文名",
        "field",
        "fieldname",
        "field_name",
        "column",
        "columnname",
        "column_name",
        "name",
    },
    "type": {
        "类型",
        "字段类型",
        "数据类型",
        "数据库类型",
        "datatype",
        "data_type",
        "type",
        "dbtype",
        "db_type",
    },
    "required": {
        "必填",
        "是否必填",
        "是否为空",
        "允许为空",
        "可为空",
        "非空",
        "notnull",
        "not_null",
        "nullable",
        "required",
        "mandatory",
    },
    "primary_key": {
        "主键",
        "是否主键",
        "pk",
        "primarykey",
        "primary_key",
        "key",
    },
    "default": {
        "默认值",
        "缺省值",
        "default",
        "defaultvalue",
        "default_value",
    },
    "description": {
        "说明",
        "字段说明",
        "描述",
        "备注",
        "注释",
        "中文名",
        "中文名称",
        "comment",
        "comments",
        "description",
        "desc",
        "remark",
        "remarks",
    },
    "length": {
        "长度",
        "字段长度",
        "精度",
        "length",
        "precision",
        "size",
    },
}

META_ALIASES = {
    "table_name": {
        "表名",
        "数据表名",
        "物理表名",
        "英文表名",
        "table",
        "table_name",
        "tablename",
    },
    "display_name": {
        "中文表名",
        "表中文名",
        "表名称",
        "业务名称",
        "逻辑表名",
        "display_name",
    },
    "description": {
        "表说明",
        "说明",
        "描述",
        "备注",
        "用途",
        "description",
        "comment",
        "remark",
    },
}

OUTPUT_COLUMNS = (
    ("index", "序号"),
    ("name", "字段名"),
    ("type", "类型"),
    ("length", "长度"),
    ("required", "必填"),
    ("primary_key", "主键"),
    ("default", "默认值"),
    ("description", "说明"),
)


@dataclass
class FieldRow:
    values: dict[str, str]


@dataclass
class TableDoc:
    table_name: str
    display_name: str = ""
    description: str = ""
    source_sheet: str = ""
    fields: list[FieldRow] = field(default_factory=list)
    filename: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    text = re.sub(r"[ \t]+", " ", text)
    return text


def normalize_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[\s:：()（）\[\]【】/\\\-_.]+", "", text)
    return text


def markdown_escape(value: Any) -> str:
    text = normalize_text(value)
    text = text.replace("\\", "\\\\").replace("|", "\\|")
    return text.replace("\n", "<br>")


def fill_merged_values(sheet) -> list[list[str]]:
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    grid = [
        [normalize_text(sheet.cell(row=row, column=col).value) for col in range(1, max_col + 1)]
        for row in range(1, max_row + 1)
    ]
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col_range, max_row_range = merged_range.bounds
        value = normalize_text(sheet.cell(row=min_row, column=min_col).value)
        if not value:
            continue
        for row in range(min_row, max_row_range + 1):
            for col in range(min_col, max_col_range + 1):
                grid[row - 1][col - 1] = value
    return [trim_row(row) for row in grid]


def trim_row(row: list[str]) -> list[str]:
    end = len(row)
    while end > 0 and not row[end - 1]:
        end -= 1
    return row[:end]


def is_empty_row(row: list[str]) -> bool:
    return not any(cell.strip() for cell in row)


def find_alias_group(value: Any, aliases: dict[str, set[str]]) -> str | None:
    key = normalize_key(value)
    if not key:
        return None
    for group, names in aliases.items():
        if key in {normalize_key(name) for name in names}:
            return group
    return None


def detect_header(row: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    seen: set[str] = set()
    for index, cell in enumerate(row):
        group = find_alias_group(cell, FIELD_HEADER_ALIASES)
        if group and group not in seen:
            mapping[index] = group
            seen.add(group)
    if "name" not in seen:
        return {}
    if len(seen) < 2:
        return {}
    return mapping


def detect_table_title(row: list[str]) -> tuple[str, str]:
    cells = [cell for cell in row if cell]
    if not cells:
        return "", ""

    joined = " ".join(cells)
    patterns = (
        r"(?:表名|数据表名|物理表名|table\s*name|table)[:：]\s*([A-Za-z0-9_\-.]+)\s*(.*)",
        r"([A-Za-z][A-Za-z0-9_\-.]{2,})\s*[\(（](.+?)[\)）]",
    )
    for pattern in patterns:
        match = re.search(pattern, joined, re.IGNORECASE)
        if match:
            table_name = normalize_text(match.group(1))
            display_name = normalize_text(match.group(2)) if len(match.groups()) > 1 else ""
            return table_name, display_name

    if len(cells) == 1 and not detect_header(row):
        value = cells[0]
        if len(value) <= 80 and any(token in value.lower() for token in ("表", "table", "_")):
            return value, ""

    return "", ""


def extract_meta(rows: list[list[str]], start: int, end: int, sheet_name: str) -> dict[str, str]:
    meta = {
        "table_name": "",
        "display_name": "",
        "description": "",
    }

    search_start = max(0, start - 8)
    for row in rows[search_start:start]:
        title_name, title_display_name = detect_table_title(row)
        if title_name and not meta["table_name"]:
            meta["table_name"] = title_name
        if title_display_name and not meta["display_name"]:
            meta["display_name"] = title_display_name

        for index, cell in enumerate(row):
            group = find_alias_group(cell, META_ALIASES)
            if not group:
                continue
            value = ""
            if index + 1 < len(row):
                value = normalize_text(row[index + 1])
            if not value:
                parts = re.split(r"[:：]", cell, maxsplit=1)
                value = normalize_text(parts[1]) if len(parts) == 2 else ""
            if value and not meta[group]:
                meta[group] = value

    if not meta["table_name"]:
        title_name, title_display_name = detect_table_title(rows[start]) if start < len(rows) else ("", "")
        meta["table_name"] = title_name
        if title_display_name:
            meta["display_name"] = title_display_name

    if not meta["table_name"]:
        meta["table_name"] = sheet_name

    if not meta["display_name"] and meta["table_name"] != sheet_name:
        meta["display_name"] = sheet_name

    return meta


def split_table_blocks(rows: list[list[str]]) -> list[tuple[int, int, dict[int, str]]]:
    headers: list[tuple[int, dict[int, str]]] = []
    for row_index, row in enumerate(rows):
        mapping = detect_header(row)
        if mapping:
            headers.append((row_index, mapping))

    blocks: list[tuple[int, int, dict[int, str]]] = []
    for header_index, (start, mapping) in enumerate(headers):
        next_header = headers[header_index + 1][0] if header_index + 1 < len(headers) else len(rows)
        end = next_header
        while end > start + 1 and is_empty_row(rows[end - 1]):
            end -= 1
        if end > start + 1:
            blocks.append((start, end, mapping))
    return blocks


def looks_like_field_name(value: str) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    if find_alias_group(text, FIELD_HEADER_ALIASES):
        return False
    return bool(re.search(r"[A-Za-z0-9_\u4e00-\u9fff]", text))


def extract_fields(rows: list[list[str]], start: int, end: int, header: dict[int, str]) -> list[FieldRow]:
    fields: list[FieldRow] = []
    empty_streak = 0
    for row in rows[start + 1:end]:
        if is_empty_row(row):
            empty_streak += 1
            if empty_streak >= 2:
                break
            continue
        empty_streak = 0

        values = {key: "" for key, _label in OUTPUT_COLUMNS}
        for column_index, group in header.items():
            if column_index < len(row):
                values[group] = normalize_text(row[column_index])

        if not looks_like_field_name(values.get("name", "")):
            continue
        fields.append(FieldRow(values=values))
    return fields


def parse_sheet(sheet) -> list[TableDoc]:
    if any(keyword in sheet.title for keyword in SKIP_SHEET_KEYWORDS):
        return []

    rows = fill_merged_values(sheet)
    blocks = split_table_blocks(rows)
    tables: list[TableDoc] = []

    for block_index, (start, end, header) in enumerate(blocks, start=1):
        meta = extract_meta(rows, start, end, sheet.title)
        fields = extract_fields(rows, start, end, header)
        if not fields:
            continue

        table_name = meta["table_name"]
        if len(blocks) > 1 and table_name == sheet.title:
            table_name = f"{sheet.title}_{block_index}"

        tables.append(
            TableDoc(
                table_name=table_name,
                display_name=meta["display_name"],
                description=meta["description"],
                source_sheet=sheet.title,
                fields=fields,
            )
        )

    return tables


def sanitize_filename(value: str, used: set[str]) -> str:
    name = normalize_text(value) or "table"
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name)
    name = re.sub(r"\s+", "_", name)
    name = name.strip("._ ")
    if not name:
        name = "table"

    candidate = f"{name}.md"
    counter = 2
    while candidate.lower() in used:
        candidate = f"{name}_{counter}.md"
        counter += 1
    used.add(candidate.lower())
    return candidate


def render_table_md(table: TableDoc) -> str:
    title = table.table_name
    lines = [
        f"# {markdown_escape(title)}",
        "",
        "## 基本信息",
        "",
        "| 项目 | 内容 |",
        "|---|---|",
        f"| 表名 | {markdown_escape(table.table_name)} |",
        f"| 中文名 | {markdown_escape(table.display_name)} |",
        f"| 来源 Sheet | {markdown_escape(table.source_sheet)} |",
        f"| 说明 | {markdown_escape(table.description)} |",
        "",
        "## 字段说明",
        "",
        "| " + " | ".join(label for _key, label in OUTPUT_COLUMNS) + " |",
        "| " + " | ".join("---" for _key, _label in OUTPUT_COLUMNS) + " |",
    ]

    for index, field_row in enumerate(table.fields, start=1):
        values = dict(field_row.values)
        if not values.get("index"):
            values["index"] = str(index)
        lines.append(
            "| "
            + " | ".join(markdown_escape(values.get(key, "")) for key, _label in OUTPUT_COLUMNS)
            + " |"
        )

    lines.append("")
    return "\n".join(lines)


def render_readme_md(tables: list[TableDoc], source: Path) -> str:
    lines = [
        "# 数据字典概览",
        "",
        f"- 来源文件：`{source.name}`",
        f"- 表数量：{len(tables)}",
        "",
        "| 序号 | 表名 | 中文名 | 字段数 | 来源 Sheet | 文档 |",
        "|---:|---|---|---:|---|---|",
    ]
    for index, table in enumerate(tables, start=1):
        lines.append(
            "| "
            + " | ".join(
                [
                    str(index),
                    markdown_escape(table.table_name),
                    markdown_escape(table.display_name),
                    str(len(table.fields)),
                    markdown_escape(table.source_sheet),
                    f"[{markdown_escape(table.filename)}]({markdown_escape(table.filename)})",
                ]
            )
            + " |"
        )
    lines.append("")
    return "\n".join(lines)


def write_text(path: Path, content: str, overwrite: bool) -> None:
    if path.exists() and not overwrite:
        raise FileExistsError(f"Refusing to overwrite existing file: {path}")
    path.write_text(content, encoding="utf-8", newline="\n")


def convert(excel_path: Path, output_dir: Path, overwrite: bool) -> list[TableDoc]:
    workbook = load_workbook(excel_path, data_only=True)
    tables: list[TableDoc] = []
    for sheet in workbook.worksheets:
        tables.extend(parse_sheet(sheet))

    if not tables:
        raise RuntimeError("No data dictionary tables were detected.")

    output_dir.mkdir(parents=True, exist_ok=True)
    used_filenames: set[str] = set()
    for table in tables:
        table.filename = sanitize_filename(table.table_name, used_filenames)
        write_text(output_dir / table.filename, render_table_md(table), overwrite=overwrite)

    write_text(output_dir / "README.md", render_readme_md(tables, excel_path), overwrite=overwrite)
    return tables


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convert an Excel data dictionary to per-table Markdown files and README.md."
    )
    parser.add_argument("excel", type=Path, help="Path to the .xlsx data dictionary.")
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output directory. Defaults to '<excel-stem>_md' next to the workbook.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing Markdown files in the output directory.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    excel_path = args.excel.resolve()
    if not excel_path.exists():
        print(f"Excel file does not exist: {excel_path}", file=sys.stderr)
        return 2
    if excel_path.suffix.lower() != ".xlsx":
        print(f"Expected a .xlsx file: {excel_path}", file=sys.stderr)
        return 2

    output_dir = args.out or excel_path.with_name(f"{excel_path.stem}_md")
    tables = convert(excel_path, output_dir.resolve(), overwrite=args.overwrite)
    print(f"Generated {len(tables)} table document(s) in: {output_dir.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
